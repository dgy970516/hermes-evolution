import asyncio
import json
import logging
import os
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any, Optional

logger = logging.getLogger("hermes.operator.db")


@dataclass
class DatabaseConfig:
    name: str
    type: str
    host: str = "localhost"
    port: int = 3306
    database: str = ""
    username: str = ""
    password: str = ""
    file_path: str = ""

    @classmethod
    def from_dict(cls, d: dict) -> "DatabaseConfig":
        return cls(
            name=d.get("name", "default"),
            type=d.get("type", "mysql"),
            host=d.get("host", "localhost"),
            port=int(d.get("port", 3306)),
            database=d.get("database", ""),
            username=d.get("username", ""),
            password=d.get("password", ""),
            file_path=d.get("file_path", ""),
        )

    @property
    def display(self) -> str:
        if self.type == "sqlite":
            return f"SQLite: {self.file_path}"
        return f"{self.type}://{self.host}:{self.port}/{self.database}"


SQL_GENERATION_PROMPT = """你是一个 SQL 专家。用户用自然语言描述了一个数据库查询需求。
请根据数据库的表结构生成对应的 SQL 查询语句。

规则：
- 只返回 SQL 语句本身，不要其他内容
- 优先使用 SELECT 查询
- 如果用户没指定表名，请根据描述猜测最可能的表
- 如果没有明确条件，默认 LIMIT 50
- 如果是统计类需求(总计、平均、分组)，生成对应的 GROUP BY 语句

用户需求：{user_input}

表名提示（如果有）：{table_hint}

数据库中的实际表名：
{available_tables}
"""

FIX_SQL_PROMPT = """之前的 SQL 查询执行失败。
错误信息：{error}
原始 SQL：{sql}
数据库可用表：{available_tables}

请分析错误原因，生成修正后的 SQL 语句。
常见问题：
- 表名不对：使用 SHOW TABLES 列出的正确表名
- 字段不对：使用 `SELECT *` 替代
- 语法问题

只返回修正后的 SQL，不要其他内容。
"""


class DatabaseClient:
    def __init__(self, configs: list[DatabaseConfig] | None = None):
        self._configs: dict[str, DatabaseConfig] = {}
        self._connections: dict[str, Any] = {}
        if configs:
            for cfg in configs:
                self._configs[cfg.name] = cfg

    def register_config(self, config: DatabaseConfig):
        self._configs[config.name] = config

    def load_configs(self, config_path: str):
        import yaml
        path = Path(config_path)
        if not path.exists():
            logger.warning(f"Database config file not found: {config_path}")
            return
        with open(path, encoding="utf-8") as f:
            data = yaml.safe_load(f)
        if data is None:
            return
        for entry in data.get("databases") or []:
            cfg = DatabaseConfig.from_dict(entry)
            self._configs[cfg.name] = cfg
        logger.info(f"Loaded {len(self._configs)} database configs from {config_path}")

    async def discover_tables(self, db_name: str) -> list[str]:
        """Get list of all tables in the database"""
        result = await self.execute_query(db_name, "SHOW TABLES", read_only=True)
        if result.get("success"):
            rows = result["data"].get("rows", [])
            return [str(row[0]) for row in rows if row]
        return []

    async def generate_sql(self, user_input: str, table_hint: str = "", llm_client=None, db_name: str = "local") -> str:
        """Use LLM to generate SQL from natural language.
        First discovers available tables to provide accurate table names."""
        if not llm_client:
            table = await self._guess_table(user_input)
            return f"SELECT * FROM `{table}` LIMIT 50"

        # Discover available tables
        available = await self.discover_tables(db_name)
        tables_str = ", ".join(available) if available else "无（查询失败）"
        if available:
            logger.info(f"Discovered {len(available)} tables: {available[:5]}...")

        prompt = SQL_GENERATION_PROMPT.format(
            user_input=user_input,
            table_hint=table_hint or "无",
            available_tables=tables_str,
        )
        try:
            sql = await llm_client.chat(prompt, f"根据需求生成 SQL: {user_input}")
            sql = sql.strip()
            if "```sql" in sql:
                sql = sql.split("```sql")[1].split("```")[0].strip()
            elif "```" in sql:
                sql = sql.split("```")[1].split("```")[0].strip()
            sql = sql.rstrip(";")
            logger.info(f"Auto-generated SQL: {sql[:100]}")
            return sql
        except Exception as e:
            logger.warning(f"SQL generation failed: {e}")
            table = await self._guess_table(user_input)
            return f"SELECT * FROM `{table}` LIMIT 50"

    async def execute_with_retry(self, db_name: str, sql: str, llm_client=None, max_retries: int = 2) -> dict:
        """Execute query with automatic error recovery"""
        result = await self.execute_query(db_name, sql)

        # 成功 → 直接返回
        if result.get("success"):
            return result

        error_msg = result.get("message", "")

        # 表不存在 → 尝试修复
        if max_retries > 0 and llm_client and ("doesn't exist" in error_msg or "does not exist" in error_msg or "1146" in error_msg):
            available = await self.discover_tables(db_name)
            if available:
                tables_str = ", ".join(available[:30])
                logger.info(f"Query failed, retrying with correct table name. Available: {tables_str[:80]}...")
                fix_prompt = FIX_SQL_PROMPT.format(
                    error=error_msg[:200],
                    sql=sql[:200],
                    available_tables=tables_str,
                )
                try:
                    fixed_sql = await llm_client.chat(fix_prompt, f"修复 SQL 错误: {error_msg[:100]}")
                    fixed_sql = fixed_sql.strip()
                    if "```sql" in fixed_sql:
                        fixed_sql = fixed_sql.split("```sql")[1].split("```")[0].strip()
                    elif "```" in fixed_sql:
                        fixed_sql = fixed_sql.split("```")[1].split("```")[0].strip()
                    fixed_sql = fixed_sql.rstrip(";")
                    logger.info(f"Fixed SQL: {fixed_sql[:100]}")
                    if fixed_sql and fixed_sql != sql:
                        result2 = await self.execute_query(db_name, fixed_sql)
                        if result2.get("success"):
                            result2["note"] = f"自动修复：原表名不可用，已使用 {_extract_table(fixed_sql)}"
                            return result2
                        # 递归重试
                        return await self.execute_with_retry(db_name, fixed_sql, llm_client, max_retries - 1)
                except Exception as e:
                    logger.warning(f"SQL fix failed: {e}")

        return result

    async def _guess_table(self, text: str) -> str:
        """Extract possible table name from user text"""
        import re
        # Common patterns: "查一下 xx 表", "xx表的数据", "sys_user"
        patterns = [
            r"查\s*.+?[`'\"](.+?)[`'\"]",  # 查一下 `sys_user`
            r"查\s*.+?(\w+?)\s*表",          # 查一下 sys_user 表
            r"(\w+?_table)",                  # sys_user_table
            r"(\w+?)\s*表的数据",             # sys_user 表的数据
            r"(\w+?)\s+表",                   # sys_user 表
        ]
        for p in patterns:
            m = re.search(p, text)
            if m:
                return m.group(1)
        return "users"  # fallback

    async def execute_query(self, db_name: str, sql: str, read_only: bool = True) -> dict:
        config = self._configs.get(db_name)
        if not config:
            available = list(self._configs.keys())
            return {"success": False, "message": f"数据库 '{db_name}' 未配置。可用: {available}"}

        sql_upper = sql.strip().upper()
        if read_only and not sql_upper.startswith(("SELECT", "SHOW", "DESC", "EXPLAIN", "WITH")):
            return {"success": False, "message": "安全模式：仅允许查询操作"}

        try:
            result = await self._do_query(config, sql)
            return {"success": True, "data": result}
        except Exception as e:
            return {"success": False, "message": f"查询失败: {e}"}

    async def export_to_excel(self, data: dict, filename: str = "query_result.xlsx") -> dict:
        """Convert query results to Excel file"""
        try:
            import openpyxl
            from openpyxl.utils import get_column_letter
        except ImportError:
            return {"success": False, "message": "请先安装 openpyxl: pip install openpyxl"}

        columns = data.get("columns", [])
        rows = data.get("rows", [])
        if not columns:
            return {"success": False, "message": "无数据可导出"}

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Query Result"

        # Header
        for col_idx, col_name in enumerate(columns, 1):
            cell = ws.cell(row=1, column=col_idx, value=col_name)
            cell.font = openpyxl.styles.Font(bold=True)

        # Data
        for row_idx, row_data in enumerate(rows, 2):
            for col_idx, val in enumerate(row_data, 1):
                ws.cell(row=row_idx, column=col_idx, value=val)

        # Auto-fit column widths
        for col_idx, col_name in enumerate(columns, 1):
            ws.column_dimensions[get_column_letter(col_idx)].width = min(len(str(col_name)) + 4, 40)

        # Save to temp file
        export_dir = Path("data/exports")
        export_dir.mkdir(parents=True, exist_ok=True)
        filepath = export_dir / filename
        wb.save(str(filepath))
        logger.info(f"Excel exported: {filepath} ({len(rows)} rows)")

        return {
            "success": True,
            "filepath": str(filepath),
            "filename": filename,
            "row_count": len(rows),
        }

    async def _do_query(self, config: DatabaseConfig, sql: str) -> dict:
        if config.type == "sqlite":
            return await self._query_sqlite(config, sql)
        elif config.type == "mysql":
            return await self._query_mysql(config, sql)
        elif config.type == "postgresql":
            return await self._query_postgres(config, sql)
        else:
            raise ValueError(f"不支持的数据库类型: {config.type}")

    async def _query_sqlite(self, config: DatabaseConfig, sql: str) -> dict:
        import aiosqlite
        db_path = config.file_path
        if not db_path:
            raise ValueError("SQLite 需要指定 file_path")
        async with aiosqlite.connect(db_path) as db:
            db.row_factory = aiosqlite.Row
            cursor = await db.execute(sql)
            rows = await cursor.fetchall()
            columns = [desc[0] for desc in cursor.description] if cursor.description else []
            return {"columns": columns, "rows": [list(row) for row in rows], "row_count": len(rows)}

    async def _query_mysql(self, config: DatabaseConfig, sql: str) -> dict:
        try:
            import asyncmy
        except ImportError:
            raise ImportError("MySQL 支持需要安装 asyncmy: pip install asyncmy")
        conn = await asyncmy.connect(
            host=config.host, port=config.port,
            user=config.username, password=config.password,
            database=config.database,
        )
        try:
            async with conn.cursor() as cur:
                await cur.execute(sql)
                rows = await cur.fetchall()
                columns = [desc[0] for desc in cur.description] if cur.description else []
                return {"columns": columns, "rows": [list(row) for row in rows], "row_count": len(rows)}
        finally:
            await conn.ensure_closed()

    async def _query_postgres(self, config: DatabaseConfig, sql: str) -> dict:
        try:
            import asyncpg
        except ImportError:
            raise ImportError("PostgreSQL 支持需要安装 asyncpg: pip install asyncpg")
        conn = await asyncpg.connect(
            host=config.host, port=config.port,
            user=config.username, password=config.password,
            database=config.database,
        )
        try:
            records = await conn.fetch(sql)
            columns = list(records[0].keys()) if records else []
            rows = [list(record.values()) for record in records]
            return {"columns": columns, "rows": rows, "row_count": len(rows)}
        finally:
            await conn.close()

    def list_databases(self) -> list[dict]:
        return [
            {"name": name, "type": cfg.type, "display": cfg.display}
            for name, cfg in self._configs.items()
        ]


def _extract_table(sql: str) -> str:
    """Extract table name from SQL"""
    import re
    m = re.search(r"FROM\s+[`'\"]?(\w+)[`'\"]?", sql, re.IGNORECASE)
    return m.group(1) if m else "?"
