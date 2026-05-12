"""
Excel Export Skill — 数据导出为 Excel 表格
============================================
用法: "导出 zsadmin 项目数据为 Excel" / "生成代码统计表"
"""
import logging
from pathlib import Path

from src.skills.base import Skill, SkillContext

logger = logging.getLogger("hermes.skills.excel")


class ExcelExportSkill(Skill):
    name = "excel_export"
    description = "导出项目数据、代码统计等为 Excel 表格"
    intents = []
    triggers = ["excel", "导出", "表格", "xlsx", "统计表", "电子表格"]

    async def can_handle(self, ctx: SkillContext) -> bool:
        text = ctx.text.lower()
        return any(t in text for t in self.triggers)

    async def execute(self, ctx: SkillContext) -> None:
        hermes = ctx.hermes

        # 匹配项目（如果有）
        proj, reason = await hermes.matcher.match(ctx.text)
        if not proj:
            yield f"⚠️ {reason}"
            return

        project_path = Path(proj.path)
        if not project_path.exists():
            yield f"❌ 项目目录不存在: {proj.path}"
            return

        yield f"📁 项目: {proj.name}"
        yield "📊 正在采集数据..."

        try:
            import openpyxl
            from openpyxl.styles import Font, Alignment, PatternFill
        except ImportError:
            yield "📦 安装 openpyxl..."
            await hermes.upgrader.ensure_dependencies(["openpyxl"])
            import openpyxl
            from openpyxl.styles import Font, Alignment, PatternFill

        wb = openpyxl.Workbook()

        # Sheet 1: 文件清单
        ws1 = wb.active
        ws1.title = "项目文件"
        headers = ["文件路径", "类型", "大小(KB)", "行数"]
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        for col, h in enumerate(headers, 1):
            cell = ws1.cell(row=1, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill

        row = 2
        total_size = 0
        total_lines = 0
        ext_counts = {}
        source_exts = {".py", ".js", ".ts", ".tsx", ".jsx", ".vue", ".java",
                       ".go", ".rs", ".cpp", ".c", ".cs", ".php", ".rb",
                       ".kt", ".swift", ".scala", ".m", ".h", ".sql"}

        for f in sorted(project_path.rglob("*")):
            if f.is_file() and f.suffix in source_exts:
                try:
                    lines = len(f.read_text(encoding="utf-8", errors="replace").splitlines())
                except Exception:
                    lines = 0
                size_kb = f.stat().st_size // 1024
                rel_path = str(f.relative_to(project_path))
                ws1.cell(row=row, column=1, value=rel_path)
                ws1.cell(row=row, column=2, value=f.suffix)
                ws1.cell(row=row, column=3, value=size_kb)
                ws1.cell(row=row, column=4, value=lines)
                total_size += size_kb
                total_lines += lines
                ext_counts[f.suffix] = ext_counts.get(f.suffix, 0) + 1
                row += 1

                if row > 1000:  # 限制行数
                    ws1.cell(row=row, column=1, value="... 超过1000行，截断")
                    break

        # Sheet 2: 统计总结
        ws2 = wb.create_sheet("统计汇总")
        ws2.cell(row=1, column=1, value="项目名称").font = header_font
        ws2.cell(row=1, column=2, value=proj.name).font = header_font
        ws2.cell(row=2, column=1, value="总文件数")
        ws2.cell(row=2, column=2, value=row - 2)
        ws2.cell(row=3, column=1, value="总行数")
        ws2.cell(row=3, column=2, value=total_lines)
        ws2.cell(row=4, column=1, value="总大小(KB)")
        ws2.cell(row=4, column=2, value=total_size)

        ws2.cell(row=6, column=1, value="语言分布").font = header_font
        ws2.cell(row=7, column=1, value="文件类型")
        ws2.cell(row=7, column=2, value="文件数")
        for i, (ext, count) in enumerate(sorted(ext_counts.items(), key=lambda x: x[1], reverse=True), 8):
            ws2.cell(row=i, column=1, value=ext)
            ws2.cell(row=i, column=2, value=count)

        # 保存
        export_dir = Path("data/exports")
        export_dir.mkdir(parents=True, exist_ok=True)
        filename = f"{proj.name}_代码统计.xlsx"
        filepath = export_dir / filename
        wb.save(str(filepath))

        yield f"✅ 已生成: {filename}"
        yield f"📊 {row - 2} 个文件, {total_lines} 行代码"

        # 发送到飞书
        if hermes.feishu_adapter and ctx.user_id.startswith("ou_"):
            yield "📤 正在发送到飞书..."
            result = await hermes.feishu_adapter.send_file(
                ctx.user_id, str(filepath), filename,
            )
            if result.get("success"):
                yield "✅ Excel 已发送到飞书"
            else:
                yield f"⚠️ 文件发送失败: {result.get('message', '')}"
        else:
            yield f"📁 {filepath}"
